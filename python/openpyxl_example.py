# working with excel spreadsheets

from openpyxl.utils import get_column_letter
# >>> get_column_letter(3)
# 'C'

# getting info from table in readonly mode
from openpyxl import load_workbook
wb = load_workbook(filename='spreadsheet_file.xlsx')
wb.sheetnames            # to get the tab or sheet names 
ws = wb['tab_name']
# search for a value in a cell 

def search_first(ws, val):
    for col in ws.iter_cols():
        for cell in col:
            if cell.value == val:
                return (cell.row, cell.column)
                  
def search_all(ws, val):
    result = []
    for col in ws.iter_cols():
        for cell in col:
            if cell.value == val:
                result.append((cell.row, cell.column))
    return result

def print_ws(ws):
    for row_cells in ws.iter_rows():
        row = '| '
        for cell in row_cells:
            row+= str(cell.value) + ' | '
        print(row)
        
